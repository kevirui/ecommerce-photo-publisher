from fastapi import FastAPI, File, UploadFile, Form, HTTPException
from fastapi.responses import JSONResponse, FileResponse
import os
import tempfile
import shutil
import uuid
import time
import threading
from pathlib import Path

# from services.database import SqlService
# from services.ftp import FtpService
from services.image_processor import process_image

app = FastAPI(title="Ecommerce Uploader Backend")

import configparser

# Leer configuración de appdesktop
config_path = Path(__file__).parent.parent / "appdesktop" / "config.ini"
config = configparser.ConfigParser(interpolation=None)

# Valores por defecto
SQL_SERVER = "tu_servidor"
SQL_DB = "tu_basededatos"
SQL_USER = "tu_usuario"
SQL_PASS = "tu_password"

FTP_HOST = "ftp.tudominio.com"
FTP_USER = "tu_ftp_user"
FTP_PASS = "tu_ftp_pass"
FTP_PATH = "/public_html/imagenes"

if config_path.exists():
    try:
        config.read(config_path, encoding="utf-8")
        if "SQL" in config:
            SQL_SERVER = config.get("SQL", "server", fallback=SQL_SERVER)
            SQL_DB = config.get("SQL", "database", fallback=SQL_DB)
            SQL_USER = config.get("SQL", "username", fallback=SQL_USER)
            SQL_PASS = config.get("SQL", "password", fallback=SQL_PASS)
            
        if "FTP" in config:
            FTP_HOST = config.get("FTP", "host", fallback=FTP_HOST)
            FTP_USER = config.get("FTP", "username", fallback=FTP_USER)
            FTP_PASS = config.get("FTP", "password", fallback=FTP_PASS)
            FTP_PATH = config.get("FTP", "remote_path", fallback=FTP_PATH)
        
        print(f"Configuración cargada exitosamente desde {config_path}")
    except Exception as e:
        print(f"Error al leer config.ini: {e}")
else:
    print(f"Advertencia: No se encontró archivo de configuración en {config_path}")

# Directorio de respaldo local
BACKUP_DIR = Path("Respaldo_Imagenes")

# Directorio de previews temporales
PREVIEWS_DIR = Path("Previews_Temp")
PREVIEWS_DIR.mkdir(parents=True, exist_ok=True)

NO_STOCK_EXCEL_PATH = Path(__file__).parent / "productos_sin_stock.xlsx"

def get_no_stock_codes() -> set:
    import openpyxl
    codes = set()
    if not NO_STOCK_EXCEL_PATH.exists():
        return codes
    try:
        workbook = openpyxl.load_workbook(str(NO_STOCK_EXCEL_PATH), read_only=True)
        sheet = workbook.active
        if sheet is not None:
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if row and row[0]:
                    codes.add(str(row[0]).strip().upper())
    except Exception as e:
        print(f"Error reading no-stock excel: {e}")
    return codes

def add_to_no_stock_excel(code: str):
    import openpyxl
    from datetime import datetime
    code = code.strip().upper()

    # Query DB to get the description and current stock
    from services.database import SqlService
    sql = SqlService(SQL_SERVER, SQL_DB, SQL_USER, SQL_PASS)
    stock_val = 0.0
    description_val = ""
    try:
        sql.connect()
        results = sql.execute("SELECT DESCRIP_ARTI, CANT_STOCK FROM ARTICULOS WHERE COD_ARTICULO = ?", (code,))
        if results:
            stock_val = float(results[0].get("CANT_STOCK", 0.0)) if results[0].get("CANT_STOCK") is not None else 0.0
            description_val = str(results[0].get("DESCRIP_ARTI", "")).strip()
    except Exception as e:
        print(f"Error fetching product data for no-stock entry {code}: {e}")
    finally:
        sql.disconnect()
    
    headers = ["Código", "Descripción", "Fecha", "Stock Fantasma"]
    
    if NO_STOCK_EXCEL_PATH.exists():
        try:
            workbook = openpyxl.load_workbook(str(NO_STOCK_EXCEL_PATH))
            sheet = workbook.active
            # Ensure headers are correct
            sheet.cell(row=1, column=1, value="Código")
            sheet.cell(row=1, column=2, value="Descripción")
            sheet.cell(row=1, column=3, value="Fecha")
            sheet.cell(row=1, column=4, value="Stock Fantasma")
        except Exception:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.append(headers)
    else:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(headers)
        
    exists = False
    # Check if code already exists in column 1
    for row_idx in range(2, sheet.max_row + 1):
        cell_val = sheet.cell(row=row_idx, column=1).value
        if cell_val and str(cell_val).strip().upper() == code:
            exists = True
            # Update description, date and stock value
            sheet.cell(row=row_idx, column=2, value=description_val)
            sheet.cell(row=row_idx, column=3, value=datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
            sheet.cell(row=row_idx, column=4, value=stock_val)
            workbook.save(str(NO_STOCK_EXCEL_PATH))
            break
            
    if not exists:
        sheet.append([code, description_val, datetime.now().strftime("%Y-%m-%d %H:%M:%S"), stock_val])
        workbook.save(str(NO_STOCK_EXCEL_PATH))

def backfill_existing_stock_in_excel():
    """Populates the Description and Stock Fantasma columns for all products already in the Excel."""
    import openpyxl
    if not NO_STOCK_EXCEL_PATH.exists():
        return
    try:
        workbook = openpyxl.load_workbook(str(NO_STOCK_EXCEL_PATH))
        sheet = workbook.active
        if sheet is None:
            return
        
        # Ensure we have all headers in place
        headers = [cell.value for cell in sheet[1]]
        
        # Detect if we need to insert 'Descripción' column
        has_desc = False
        desc_col_idx = 2
        for idx, h in enumerate(headers):
            if h and "descrip" in str(h).lower():
                has_desc = True
                desc_col_idx = idx + 1
                break
                
        # If 'Descripción' header is missing, we re-format the header row to:
        # Código | Descripción | Fecha | Stock Fantasma
        # And shift columns if needed, or simply force the structure
        sheet.cell(row=1, column=1, value="Código")
        sheet.cell(row=1, column=2, value="Descripción")
        sheet.cell(row=1, column=3, value="Fecha")
        sheet.cell(row=1, column=4, value="Stock Fantasma")
        
        from services.database import SqlService
        sql = SqlService(SQL_SERVER, SQL_DB, SQL_USER, SQL_PASS)
        sql.connect()
        
        updated = False
        
        # Iteramos cada fila. En la estructura original, la columna 2 era "Fecha" y la columna 3 era "Stock Fantasma".
        # Vamos a leer los códigos de la columna 1, buscar descripción y stock en BD,
        # y re-escribir de forma limpia la fila.
        for r_idx in range(2, sheet.max_row + 1):
            code_cell = sheet.cell(row=r_idx, column=1)
            if code_cell.value:
                code = str(code_cell.value).strip().upper()
                try:
                    results = sql.execute("SELECT DESCRIP_ARTI, CANT_STOCK FROM ARTICULOS WHERE COD_ARTICULO = ?", (code,))
                    if results:
                        desc_val = str(results[0].get("DESCRIP_ARTI", "")).strip()
                        stock_val = float(results[0].get("CANT_STOCK", 0.0)) if results[0].get("CANT_STOCK") is not None else 0.0
                        
                        # Si la fecha estaba en la columna 2, la pasamos a la columna 3
                        # Si no hay fecha, ponemos la actual
                        old_date = sheet.cell(row=r_idx, column=2).value
                        if old_date and ("-" in str(old_date) or "/" in str(old_date)):
                            date_val = old_date
                        else:
                            date_val = sheet.cell(row=r_idx, column=3).value or ""
                        
                        sheet.cell(row=r_idx, column=2, value=desc_val)
                        sheet.cell(row=r_idx, column=3, value=date_val)
                        sheet.cell(row=r_idx, column=4, value=stock_val)
                        updated = True
                except Exception as e:
                    print(f"Error querying data for backfill code {code}: {e}")
        
        sql.disconnect()
        if updated:
            workbook.save(str(NO_STOCK_EXCEL_PATH))
            print("Backfilled existing descriptions and stocks in excel successfully.")
    except Exception as e:
        print(f"Error during backfilling stock in excel: {e}")

# Run backfill when backend is loaded/imported
backfill_existing_stock_in_excel()

# Almacén en memoria de previews activos: {preview_id: {path, article_code, created_at}}
_preview_store: dict[str, dict] = {}
_preview_lock = threading.Lock()

PREVIEW_EXPIRY_SECONDS = 30 * 60  # 30 minutos


def _cleanup_stale_previews():
    """Elimina previews temporales que tengan más de 30 minutos."""
    while True:
        time.sleep(300)  # Revisar cada 5 minutos
        now = time.time()
        to_remove = []
        with _preview_lock:
            for pid, info in _preview_store.items():
                if now - info["created_at"] > PREVIEW_EXPIRY_SECONDS:
                    to_remove.append(pid)
            for pid in to_remove:
                info = _preview_store.pop(pid, None)
                if info:
                    if Path(info["path"]).exists():
                        try:
                            os.remove(info["path"])
                        except Exception:
                            pass
                    if "backup_path" in info and Path(info["backup_path"]).exists():
                        try:
                            os.remove(info["backup_path"])
                        except Exception:
                            pass
        if to_remove:
            print(f"Limpieza: {len(to_remove)} preview(s) expirado(s) eliminados.")


# Iniciar hilo de limpieza como daemon
_cleanup_thread = threading.Thread(target=_cleanup_stale_previews, daemon=True)
_cleanup_thread.start()

# ==========================================
# Endpoint: Preview (Paso 1 - Procesar con IA)
# ==========================================
@app.post("/api/v1/photos/preview")
async def preview_photo(
    file: UploadFile = File(...),
    article_code: str = Form(...),
    include_stamp: bool = Form(False),
    watermark_opacity: float = Form(0.05)
):
    """Procesa la imagen con IA y devuelve un preview_id para visualizarla."""
    # Validación en Base de Datos
    from services.database import SqlService
    sql = SqlService(SQL_SERVER, SQL_DB, SQL_USER, SQL_PASS)
    try:
        sql.connect()
        results = sql.execute("SELECT * FROM ARTICULOS WHERE COD_ARTICULO = ?", (article_code,))
        if not results:
            return JSONResponse(status_code=404, content={"error": "Artículo no encontrado en la base de datos."})
    except Exception as e:
        return JSONResponse(status_code=500, content={"error": f"Error de Base de Datos (Validación): {e}"})
    finally:
        sql.disconnect()

    # Guardar archivo subido temporalmente y procesar
    with tempfile.TemporaryDirectory() as tmpdir:
        input_path = Path(tmpdir) / file.filename
        with open(input_path, "wb") as buffer:
            buffer.write(await file.read())

        try:
            processed_path = process_image(input_path, include_stamp=include_stamp, watermark_opacity=watermark_opacity)
            backup_temp_path = process_image(input_path, include_stamp=False, watermark_opacity=0.0, include_watermark=False)
        except Exception as e:
            return JSONResponse(status_code=500, content={"error": f"Error procesando imagen: {e}"})

        # Copiar la imagen procesada a carpeta de previews con ID único
        preview_id = str(uuid.uuid4())
        preview_filename = f"{preview_id}.jpg"
        preview_dest = PREVIEWS_DIR / preview_filename
        shutil.copy2(processed_path, preview_dest)

        # Copiar la versión limpia para el respaldo local a la carpeta de previews
        backup_filename = f"{preview_id}_backup.jpg"
        backup_dest = PREVIEWS_DIR / backup_filename
        shutil.copy2(backup_temp_path, backup_dest)

        try:
            os.remove(processed_path)
            os.remove(backup_temp_path)
        except Exception:
            pass

    # Registrar en el store
    with _preview_lock:
        _preview_store[preview_id] = {
            "path": str(preview_dest),
            "backup_path": str(backup_dest),
            "article_code": article_code,
            "created_at": time.time(),
        }

    return {
        "preview_id": preview_id,
        "preview_url": f"/api/v1/photos/preview/{preview_id}",
        "message": "Imagen procesada con IA. Revisa el preview antes de confirmar."
    }


# ==========================================
# Endpoint: Servir imagen de preview
# ==========================================
@app.get("/api/v1/photos/preview/{preview_id}")
async def get_preview_image(preview_id: str):
    """Sirve la imagen procesada temporalmente para visualización."""
    with _preview_lock:
        info = _preview_store.get(preview_id)

    if not info:
        raise HTTPException(status_code=404, detail="Preview no encontrado o expirado.")

    preview_path = Path(info["path"])
    if not preview_path.exists():
        with _preview_lock:
            _preview_store.pop(preview_id, None)
        raise HTTPException(status_code=404, detail="Archivo de preview no encontrado.")

    return FileResponse(str(preview_path), media_type="image/jpeg")


# ==========================================
# Endpoint: Confirm (Paso 2 - Subir a FTP/BD)
# ==========================================
@app.post("/api/v1/photos/confirm")
async def confirm_upload(
    preview_id: str = Form(...),
    article_code: str = Form(...),
    image_index: int = Form(0),
):
    """Confirma la subida de un preview ya procesado: sube a FTP y actualiza BD."""
    # Validación de existencia en Base de Datos
    from services.database import SqlService
    sql = SqlService(SQL_SERVER, SQL_DB, SQL_USER, SQL_PASS)
    try:
        sql.connect()
        results = sql.execute("SELECT * FROM ARTICULOS WHERE COD_ARTICULO = ?", (article_code,))
        if not results:
            return JSONResponse(status_code=404, content={"error": f"El artículo '{article_code}' no existe en la base de datos."})
    except Exception as e:
        return JSONResponse(status_code=500, content={"error": f"Error de Base de Datos (Validación): {e}"})
    finally:
        sql.disconnect()

    # Obtener info del preview
    with _preview_lock:
        info = _preview_store.get(preview_id)

    if not info:
        raise HTTPException(status_code=404, detail="Preview no encontrado o expirado. Vuelve a procesar la imagen.")

    processed_path = Path(info["path"])
    backup_temp_path = Path(info.get("backup_path", info["path"]))
    if not processed_path.exists():
        with _preview_lock:
            _preview_store.pop(preview_id, None)
        raise HTTPException(status_code=404, detail="Archivo de preview no encontrado.")

    file_name = f"{article_code}.jpg" if image_index == 0 else f"{article_code}_{image_index}.jpg"

    # Guardar copia en Respaldo Local (versión sin marca de agua)
    BACKUP_DIR.mkdir(parents=True, exist_ok=True)
    backup_file_path = BACKUP_DIR / file_name
    try:
        shutil.copy2(backup_temp_path, backup_file_path)
    except Exception as e:
        print(f"Advertencia: No se pudo crear el respaldo local: {e}")

    # Subida al Servidor FTP
    from services.ftp import FtpService
    ftp = FtpService(FTP_HOST, port=21, username=FTP_USER, password=FTP_PASS, remote_path=FTP_PATH)
    try:
        ftp.connect()
        ftp.upload_file(processed_path, file_name)
    except Exception as e:
        return JSONResponse(status_code=500, content={"error": f"Error subiendo a FTP: {e}"})
    finally:
        ftp.disconnect()

    # Actualización en la Base de Datos
    from services.database import SqlService
    sql = SqlService(SQL_SERVER, SQL_DB, SQL_USER, SQL_PASS)
    try:
        sql.connect()
        if image_index == 0:
            sql.call_procedure(
                "eco_articulos_publi_web_actua",
                {
                    "cod_articulo": article_code,
                    "web_publi": "S",
                    "web_imagen": file_name,
                }
            )
        else:
            sql.call_procedure(
                "eco_articulos_imagenes_actua",
                {
                    "cod_articulo": article_code,
                    "web_imagen": file_name,
                    "indice": f"_{image_index}",
                }
            )
    except Exception as e:
        return JSONResponse(status_code=500, content={"error": f"Error ejecutando SP en BD: {e}"})
    finally:
        sql.disconnect()

    # Limpiar el preview temporal
    with _preview_lock:
        _preview_store.pop(preview_id, None)
    try:
        os.remove(processed_path)
    except Exception:
        pass
    try:
        if backup_temp_path != processed_path and backup_temp_path.exists():
            os.remove(backup_temp_path)
    except Exception:
        pass

    return {
        "message": "Imagen confirmada, subida por FTP y registrada en BD con éxito",
        "article": article_code,
        "image_index": image_index,
        "backup_path": str(backup_file_path)
    }


# ==========================================
# Endpoint Legacy: Upload directo (compatibilidad)
# ==========================================
@app.post("/api/v1/photos/upload")
async def upload_photo(
    file: UploadFile = File(...),
    article_code: str = Form(...),
    include_stamp: bool = Form(False),
    image_index: int = Form(0),
    watermark_opacity: float = Form(0.3)
):
    file_name = f"{article_code}.jpg" if image_index == 0 else f"{article_code}_{image_index}.jpg"

    # ==========================================
    # 1. Validación en Base de Datos
    # ==========================================
    from services.database import SqlService
    sql = SqlService(SQL_SERVER, SQL_DB, SQL_USER, SQL_PASS)
    try:
        sql.connect()
        results = sql.execute("SELECT * FROM ARTICULOS WHERE COD_ARTICULO = ?", (article_code,))
        if not results:
            return JSONResponse(status_code=404, content={"error": "Artículo no encontrado en la base de datos."})
    except Exception as e:
        return JSONResponse(status_code=500, content={"error": f"Error de Base de Datos (Validación): {e}"})
    finally:
        # No desconectamos todavía porque usaremos sql más adelante
        pass

    # ==========================================
    # 2. Procesamiento de la Imagen (IA, Recorte)
    # ==========================================
    with tempfile.TemporaryDirectory() as tmpdir:
        input_path = Path(tmpdir) / file.filename
        
        # Guardar archivo subido temporalmente
        with open(input_path, "wb") as buffer:
            buffer.write(await file.read())

        try:
            # Procesar imagen (Quitar fondo, ajustar, poner marca de agua y sello)
            processed_path = process_image(input_path, include_stamp=include_stamp, watermark_opacity=watermark_opacity)
            backup_temp_path = process_image(input_path, include_stamp=False, watermark_opacity=0.0, include_watermark=False)
        except Exception as e:
             sql.disconnect()
             return JSONResponse(status_code=500, content={"error": f"Error procesando imagen: {e}"})

        # ==========================================
        # 3. Guardar copia en Respaldo Local (versión sin marca de agua)
        # ==========================================
        BACKUP_DIR.mkdir(parents=True, exist_ok=True)
        backup_file_path = BACKUP_DIR / file_name
        
        try:
            shutil.copy2(backup_temp_path, backup_file_path)
        except Exception as e:
            # No interrumpir el proceso si falla el respaldo, solo hacer log
            print(f"Advertencia: No se pudo crear el respaldo local: {e}")

        # ==========================================
        # 4. Subida al Servidor FTP
        # ==========================================
        from services.ftp import FtpService
        ftp = FtpService(FTP_HOST, port=21, username=FTP_USER, password=FTP_PASS, remote_path=FTP_PATH)
        try:
            ftp.connect()
            ftp.upload_file(processed_path, file_name)
        except Exception as e:
            sql.disconnect()
            try:
                os.remove(processed_path)
                os.remove(backup_temp_path)
            except Exception:
                pass
            return JSONResponse(status_code=500, content={"error": f"Error subiendo a FTP: {e}"})
        finally:
            ftp.disconnect()

        # ==========================================
        # 5. Actualización en la Base de Datos
        # ==========================================
        try:
            if image_index == 0:
                sql.call_procedure(
                    "eco_articulos_publi_web_actua",
                    {
                        "cod_articulo": article_code,
                        "web_publi": "S",
                        "web_imagen": file_name,
                    }
                )
            else:
                sql.call_procedure(
                    "eco_articulos_imagenes_actua",
                    {
                        "cod_articulo": article_code,
                        "web_imagen": file_name,
                        "indice": f"_{image_index}",
                    }
                )
        except Exception as e:
            try:
                os.remove(processed_path)
                os.remove(backup_temp_path)
            except Exception:
                pass
            return JSONResponse(status_code=500, content={"error": f"Error ejecutando SP en BD: {e}"})
        finally:
            sql.disconnect()

        # Cleanup
        try:
            os.remove(processed_path)
            os.remove(backup_temp_path)
        except Exception:
            pass

    return {
        "message": "Imagen procesada, subida por FTP y registrada en BD con éxito", 
        "article": article_code,
        "image_index": image_index,
        "backup_path": str(backup_file_path)
    }

def normalize_header(text: str) -> str:
    import unicodedata
    if not text:
        return ""
    normalized = "".join(
        c for c in unicodedata.normalize("NFD", text)
        if unicodedata.category(c) != "Mn"
    )
    return normalized.lower().strip()

@app.get("/api/v1/articles/pending")
def get_pending_articles(pending_only: bool = True):
    from services.database import SqlService
    sql = SqlService(SQL_SERVER, SQL_DB, SQL_USER, SQL_PASS)
    try:
        sql.connect()
        query = """
        SELECT
            A.COD_ARTICULO,
            A.DESCRIP_ARTI,
            A.WEB_PUBLI,
            A.WEB_IMAGEN_PROVE,
            A.CANT_STOCK,
            G1.DESCRIP_AGRU AS rubro,
            G2.DESCRIP_AGRU AS grupo,
            G3.DESCRIP_AGRU AS subgrupo,
            COUNT(I.COD_ARTICULO) AS cant_imagenes
        FROM ARTICULOS A
        LEFT JOIN ARTICULOS_IMAGENES I
            ON A.COD_ARTICULO = I.COD_ARTICULO
        LEFT JOIN AGRUPACIONES G1 ON A.AGRU_1 = G1.CODI_AGRU AND G1.NUM_AGRU = 1
        LEFT JOIN AGRUPACIONES G2 ON A.AGRU_2 = G2.CODI_AGRU AND G2.NUM_AGRU = 2
        LEFT JOIN AGRUPACIONES G3 ON A.AGRU_3 = G3.CODI_AGRU AND G3.NUM_AGRU = 3
        """
        
        if pending_only:
            query += " WHERE A.WEB_PUBLI <> 'S' OR A.WEB_IMAGEN_PROVE IS NULL OR A.WEB_IMAGEN_PROVE = '' "
            
        query += """
        GROUP BY
            A.COD_ARTICULO,
            A.DESCRIP_ARTI,
            A.WEB_PUBLI,
            A.WEB_IMAGEN_PROVE,
            A.CANT_STOCK,
            G1.DESCRIP_AGRU,
            G2.DESCRIP_AGRU,
            G3.DESCRIP_AGRU
        """
        
        no_stock_set = get_no_stock_codes()
        results = sql.execute(query)
        articles = []
        for row in results:
            code = row.get("COD_ARTICULO", "")
            if code and code.strip().upper() in no_stock_set:
                continue
            descrip = row.get("DESCRIP_ARTI", "")
            web_publi = row.get("WEB_PUBLI", "")
            imagen_prov = row.get("WEB_IMAGEN_PROVE", "")
            cant_stock = float(row.get("CANT_STOCK", 0.0)) if row.get("CANT_STOCK") is not None else 0.0
            cant_imagenes = int(row.get("cant_imagenes", 0)) if row.get("cant_imagenes") is not None else 0
            rubro = row.get("rubro", "")
            grupo = row.get("grupo", "")
            subgrupo = row.get("subgrupo", "")
            
            is_publi = (web_publi == "S")
            has_any_image = (imagen_prov is not None and str(imagen_prov).strip() != "") or (cant_imagenes > 0)
            
            if pending_only and has_any_image:
                continue
                
            if not is_publi:
                estado = "PENDIENTE"
                observaciones = "El artículo no está marcado para publicar en la web."
            elif not has_any_image:
                estado = "INCOMPLETO"
                observaciones = "Marcado para publicar pero no tiene ninguna imagen."
            else:
                estado = "PUBLICADO"
                observaciones = "Artículo publicado."
                
            if cant_stock <= 0:
                continue
                
            articles.append({
                "codigo": code.strip() if code else "",
                "descripcion": descrip.strip() if descrip else "",
                "estado": estado,
                "stock": cant_stock,
                "cant_imagenes": cant_imagenes,
                "rubro": rubro.strip() if rubro else "OTROS",
                "grupo": grupo.strip() if grupo else "OTROS",
                "subgrupo": subgrupo.strip() if subgrupo else "OTROS",
                "observaciones": observaciones
            })
            
        return {"articles": articles}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error en BD: {e}")
    finally:
        sql.disconnect()

@app.get("/api/v1/articles/categories")
def get_articles_categories(pending_only: bool = True):
    from services.database import SqlService
    sql = SqlService(SQL_SERVER, SQL_DB, SQL_USER, SQL_PASS)
    try:
        sql.connect()
        # Query that fetches only distinct Rubro, Grupo, Subgrupo of pending items in stock
        query = """
        SELECT DISTINCT
            G1.DESCRIP_AGRU AS rubro,
            G2.DESCRIP_AGRU AS grupo,
            G3.DESCRIP_AGRU AS subgrupo
        FROM ARTICULOS A
        LEFT JOIN ARTICULOS_IMAGENES I ON A.COD_ARTICULO = I.COD_ARTICULO
        LEFT JOIN AGRUPACIONES G1 ON A.AGRU_1 = G1.CODI_AGRU AND G1.NUM_AGRU = 1
        LEFT JOIN AGRUPACIONES G2 ON A.AGRU_2 = G2.CODI_AGRU AND G2.NUM_AGRU = 2
        LEFT JOIN AGRUPACIONES G3 ON A.AGRU_3 = G3.CODI_AGRU AND G3.NUM_AGRU = 3
        WHERE A.CANT_STOCK > 0
        """
        if pending_only:
            query += " AND (A.WEB_PUBLI <> 'S' OR A.WEB_IMAGEN_PROVE IS NULL OR A.WEB_IMAGEN_PROVE = '') "
            
        no_stock_set = get_no_stock_codes()
        results = sql.execute(query)
        
        # Build hierarchy tree: Rubro -> Grupo -> Subgrupo
        hierarchy = {}
        for row in results:
            rubro = (row.get("rubro") or "").strip() or "OTROS"
            grupo = (row.get("grupo") or "").strip() or "OTROS"
            subgrupo = (row.get("subgrupo") or "").strip() or "OTROS"
            
            if rubro not in hierarchy:
                hierarchy[rubro] = {}
            if grupo not in hierarchy[rubro]:
                hierarchy[rubro][grupo] = set()
            hierarchy[rubro][grupo].add(subgrupo)
            
        # Format tree as lists of dicts
        tree = []
        for r_name, g_dict in sorted(hierarchy.items()):
            grupos_list = []
            for g_name, s_set in sorted(g_dict.items()):
                grupos_list.append({
                    "nombre": g_name,
                    "subgrupos": sorted(list(s_set))
                })
            tree.append({
                "nombre": r_name,
                "grupos": grupos_list
            })
            
        return {"categories": tree}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error en BD: {e}")
    finally:
        sql.disconnect()

@app.post("/api/v1/articles/{code}/has-photo")
def mark_article_has_photo(code: str):
    """Marca un artículo como que ya tiene foto ejecutando el SP en BD con el nombre de imagen por defecto."""
    from services.database import SqlService
    sql = SqlService(SQL_SERVER, SQL_DB, SQL_USER, SQL_PASS)
    try:
        sql.connect()
        # El nombre de la imagen principal por defecto es {code}.jpg
        file_name = f"{code.strip().upper()}.jpg"
        sql.call_procedure(
            "eco_articulos_publi_web_actua",
            {
                "cod_articulo": code.strip().upper(),
                "web_publi": "S",
                "web_imagen": file_name,
            }
        )
        return {"message": "Artículo marcado como 'ya tiene foto' con éxito", "article": code}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error ejecutando SP en BD: {e}")
    finally:
        sql.disconnect()

@app.post("/api/v1/articles/{code}/no-stock")
def mark_article_no_stock(code: str):
    """Marca un artículo como 'sin stock' (stock fantasma) agregándolo al excel local."""
    try:
        add_to_no_stock_excel(code)
        return {"message": "Artículo agregado al reporte de stock fantasma", "article": code}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al escribir en excel: {e}")

@app.post("/api/v1/articles/pending/from-excel")
async def get_pending_articles_from_excel(file: UploadFile = File(...)):
    import tempfile
    import openpyxl
    import xlrd
    
    suffix = Path(file.filename).suffix.lower()
    if suffix not in (".xlsx", ".xls"):
        raise HTTPException(status_code=400, detail="El archivo debe ser un Excel (.xlsx o .xls)")

    codes = []
    
    with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
        tmp.write(await file.read())
        tmp_path = Path(tmp.name)
        
    try:
        if suffix == ".xlsx":
            workbook = openpyxl.load_workbook(filename=str(tmp_path), read_only=True, data_only=True)
            sheet = workbook.active
            if sheet is not None:
                first_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=False))
                code_col_idx = None
                accepted_names = {"codigo", "articulo", "cod_articulo", "code", "cod"}
                for cell in first_row:
                    if cell.value is None:
                        continue
                    normalized_val = normalize_header(str(cell.value))
                    normalized_val_clean = normalized_val.replace("_", "").replace("-", "")
                    if normalized_val in accepted_names or normalized_val_clean in accepted_names:
                        code_col_idx = cell.column
                        break
                        
                if code_col_idx is not None:
                    for row in sheet.iter_rows(min_row=2, values_only=False):
                        cell_val = None
                        for cell in row:
                            if cell.column == code_col_idx:
                                cell_val = cell.value
                                break
                        if cell_val is not None:
                            if isinstance(cell_val, float) and cell_val.is_integer():
                                code_str = str(int(cell_val)).strip()
                            else:
                                code_str = str(cell_val).strip()
                            if code_str and code_str not in codes:
                                codes.append(code_str)
            workbook.close()
        else:
            workbook = xlrd.open_workbook(str(tmp_path))
            sheet = workbook.sheet_by_index(0)
            first_row = [sheet.cell_value(0, col_idx) for col_idx in range(sheet.ncols)]
            code_col_idx = None
            accepted_names = {"codigo", "articulo", "cod_articulo", "code", "cod"}
            for col_idx, cell_val in enumerate(first_row):
                if cell_val is None:
                    continue
                normalized_val = normalize_header(str(cell_val))
                normalized_val_clean = normalized_val.replace("_", "").replace("-", "")
                if normalized_val in accepted_names or normalized_val_clean in accepted_names:
                    code_col_idx = col_idx
                    break
                    
            if code_col_idx is not None:
                for row_idx in range(1, sheet.nrows):
                    cell_val = sheet.cell_value(row_idx, code_col_idx)
                    if cell_val is not None:
                        if isinstance(cell_val, float) and cell_val.is_integer():
                            code_str = str(int(cell_val)).strip()
                        else:
                            code_str = str(cell_val).strip()
                        if code_str and code_str not in codes:
                            codes.append(code_str)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error leyendo el archivo Excel: {e}")
    finally:
        if tmp_path.exists():
            tmp_path.unlink()

    if not codes:
        return {"articles": []}

    from services.database import SqlService
    sql = SqlService(SQL_SERVER, SQL_DB, SQL_USER, SQL_PASS)
    try:
        sql.connect()
        articles = []
        chunk_size = 500
        for i in range(0, len(codes), chunk_size):
            chunk = codes[i:i + chunk_size]
            placeholders = ",".join("?" for _ in chunk)
            query = f"""
            SELECT
                A.COD_ARTICULO,
                A.DESCRIP_ARTI,
                A.WEB_PUBLI,
                A.WEB_IMAGEN_PROVE,
                A.CANT_STOCK,
                G1.DESCRIP_AGRU AS rubro,
                G2.DESCRIP_AGRU AS grupo,
                G3.DESCRIP_AGRU AS subgrupo,
                COUNT(I.COD_ARTICULO) AS cant_imagenes
            FROM ARTICULOS A
            LEFT JOIN ARTICULOS_IMAGENES I
                ON A.COD_ARTICULO = I.COD_ARTICULO
            LEFT JOIN AGRUPACIONES G1 ON A.AGRU_1 = G1.CODI_AGRU AND G1.NUM_AGRU = 1
            LEFT JOIN AGRUPACIONES G2 ON A.AGRU_2 = G2.CODI_AGRU AND G2.NUM_AGRU = 2
            LEFT JOIN AGRUPACIONES G3 ON A.AGRU_3 = G3.CODI_AGRU AND G3.NUM_AGRU = 3
            WHERE A.COD_ARTICULO IN ({placeholders})
            GROUP BY
                A.COD_ARTICULO,
                A.DESCRIP_ARTI,
                A.WEB_PUBLI,
                A.WEB_IMAGEN_PROVE,
                A.CANT_STOCK,
                G1.DESCRIP_AGRU,
                G2.DESCRIP_AGRU,
                G3.DESCRIP_AGRU
            """
            no_stock_set = get_no_stock_codes()
            results = sql.execute(query, tuple(chunk))
            
            for row in results:
                code = row.get("COD_ARTICULO", "")
                if code and code.strip().upper() in no_stock_set:
                    continue
                descrip = row.get("DESCRIP_ARTI", "")
                web_publi = row.get("WEB_PUBLI", "")
                imagen_prov = row.get("WEB_IMAGEN_PROVE", "")
                cant_stock = float(row.get("CANT_STOCK", 0.0)) if row.get("CANT_STOCK") is not None else 0.0
                cant_imagenes = int(row.get("cant_imagenes", 0)) if row.get("cant_imagenes") is not None else 0
                rubro = row.get("rubro", "")
                grupo = row.get("grupo", "")
                subgrupo = row.get("subgrupo", "")
                
                is_publi = (web_publi == "S")
                has_any_image = (imagen_prov is not None and str(imagen_prov).strip() != "") or (cant_imagenes > 0)
                
                if is_publi and has_any_image:
                    continue
                    
                if not is_publi:
                    estado = "PENDIENTE"
                    observaciones = "El artículo no está marcado para publicar en la web."
                else:
                    estado = "INCOMPLETO"
                    observaciones = "Marcado para publicar pero no tiene ninguna imagen."
                    
                if cant_stock <= 0:
                    continue
                    
                articles.append({
                    "codigo": code.strip() if code else "",
                    "descripcion": descrip.strip() if descrip else "",
                    "estado": estado,
                    "stock": cant_stock,
                    "cant_imagenes": cant_imagenes,
                    "rubro": rubro.strip() if rubro else "OTROS",
                    "grupo": grupo.strip() if grupo else "OTROS",
                    "subgrupo": subgrupo.strip() if subgrupo else "OTROS",
                    "observaciones": observaciones
                })
                
        return {"articles": articles}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error de BD: {e}")
    finally:
        sql.disconnect()

if __name__ == "__main__":
    import uvicorn
    uvicorn.run("main:app", host="0.0.0.0", port=8000, reload=True)
