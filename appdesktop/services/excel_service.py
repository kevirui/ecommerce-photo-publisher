"""
Servicio de lectura de archivos Excel para artículos.

Lee un archivo Excel opcional con columnas Código, Descripción y
Cantidad Fotos. Si no se proporciona un archivo Excel, la aplicación
funciona normalmente detectando artículos por nombre de imagen.
"""

from __future__ import annotations

import logging
from pathlib import Path
from typing import Any, Optional

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from models.article import Article

logger = logging.getLogger(__name__)

# Nombres de columna esperados (case-insensitive)
COLUMN_CODIGO = "código"
COLUMN_CODIGO_ALT = "codigo"
COLUMN_DESCRIPCION = "descripción"
COLUMN_DESCRIPCION_ALT = "descripcion"
COLUMN_CANTIDAD = "cantidad fotos"
COLUMN_CANTIDAD_ALT = "cantidad"


class ExcelServiceError(Exception):
    """Excepción base para errores del servicio Excel."""
    pass


class ExcelService:
    """
    Servicio para leer datos de artículos desde un archivo Excel.

    Lee un archivo Excel con las columnas: Código, Descripción,
    Cantidad Fotos. Funciona incluso sin un archivo Excel, en cuyo caso
    los artículos se detectan automáticamente por nombre de imagen.

    Attributes:
        file_path: Ruta al archivo Excel, o None si no se usa.

    Example:
        >>> excel = ExcelService(Path("articulos.xlsx"))
        >>> excel.load()
        True
        >>> data = excel.get_articles()
        >>> data["R123"]["description"]
        'Tornillo hexagonal 10mm'
    """

    def __init__(self, file_path: Optional[Path] = None) -> None:
        """
        Inicializa el servicio Excel.

        Args:
            file_path: Ruta al archivo Excel. None si no se usa Excel.
        """
        self._file_path = Path(file_path) if file_path else None
        self._articles_data: dict[str, dict[str, Any]] = {}
        self._loaded = False
        self._column_map: dict[str, int] = {}

    @property
    def has_data(self) -> bool:
        """Indica si hay datos cargados desde un Excel."""
        return self._loaded and len(self._articles_data) > 0

    @property
    def file_path(self) -> Optional[Path]:
        """Ruta al archivo Excel."""
        return self._file_path

    @property
    def article_count(self) -> int:
        """Cantidad de artículos cargados desde el Excel."""
        return len(self._articles_data)

    # ================================================================
    # Carga del archivo
    # ================================================================

    def load(self) -> bool:
        """
        Carga y parsea el archivo Excel.

        Lee la primera hoja del workbook, detecta las columnas por nombre
        y carga los datos de cada fila en un diccionario interno.

        Returns:
            True si la carga fue exitosa. False si no hay archivo configurado
            o si ocurre un error.
        """
        if self._file_path is None:
            logger.info("No se configuró archivo Excel. Modo sin Excel activo.")
            return False

        if not self._file_path.exists():
            logger.error(f"Archivo Excel no encontrado: {self._file_path}")
            return False

        try:
            logger.info(f"Cargando archivo Excel: {self._file_path}")

            workbook = load_workbook(
                filename=str(self._file_path),
                read_only=True,
                data_only=True,
            )
            sheet: Worksheet = workbook.active

            if sheet is None:
                logger.error("El archivo Excel no tiene hojas activas.")
                workbook.close()
                return False

            # Detectar columnas por nombre en la primera fila
            if not self._detect_columns(sheet):
                workbook.close()
                return False

            # Leer datos de cada fila
            self._articles_data.clear()
            row_count = 0

            for row in sheet.iter_rows(min_row=2, values_only=False):
                row_data = {cell.column: cell.value for cell in row}
                code = self._get_cell_value(row_data, "codigo")

                if code is None or str(code).strip() == "":
                    continue  # Fila vacía o sin código

                code_str = str(code).strip().upper()
                description = self._get_cell_value(row_data, "descripcion") or ""
                quantity = self._get_cell_value(row_data, "cantidad")

                # Parsear cantidad como entero
                qty_int = 0
                if quantity is not None:
                    try:
                        qty_int = int(quantity)
                    except (ValueError, TypeError):
                        qty_int = 0

                self._articles_data[code_str] = {
                    "description": str(description).strip(),
                    "quantity": qty_int,
                }
                row_count += 1

            workbook.close()

            self._loaded = True
            logger.info(
                f"Excel cargado exitosamente: {row_count} artículos leídos "
                f"desde '{self._file_path.name}'"
            )
            return True

        except Exception as e:
            logger.error(f"Error al cargar archivo Excel: {e}", exc_info=True)
            self._loaded = False
            return False

    def _detect_columns(self, sheet: Worksheet) -> bool:
        """
        Detecta las posiciones de las columnas esperadas en la primera fila.

        Args:
            sheet: Hoja de trabajo de openpyxl.

        Returns:
            True si al menos la columna 'Código' fue encontrada.
        """
        self._column_map.clear()

        first_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=False))

        for cell in first_row:
            if cell.value is None:
                continue

            header = str(cell.value).strip().lower()

            if header in (COLUMN_CODIGO, COLUMN_CODIGO_ALT, "code", "cod"):
                self._column_map["codigo"] = cell.column
            elif header in (COLUMN_DESCRIPCION, COLUMN_DESCRIPCION_ALT,
                            "description", "desc"):
                self._column_map["descripcion"] = cell.column
            elif header in (COLUMN_CANTIDAD, COLUMN_CANTIDAD_ALT,
                            "cant", "fotos", "qty"):
                self._column_map["cantidad"] = cell.column

        if "codigo" not in self._column_map:
            logger.error(
                "No se encontró la columna 'Código' en el archivo Excel. "
                "Columnas detectadas: "
                f"{[str(c.value) for c in first_row if c.value]}"
            )
            return False

        logger.debug(f"Columnas Excel detectadas: {self._column_map}")
        return True

    def _get_cell_value(
        self,
        row_data: dict[int, Any],
        column_key: str,
    ) -> Optional[Any]:
        """
        Obtiene el valor de una celda por clave de columna mapeada.

        Args:
            row_data: Diccionario {columna: valor} de la fila.
            column_key: Clave de la columna en el mapa interno.

        Returns:
            Valor de la celda, o None si la columna no existe.
        """
        col_index = self._column_map.get(column_key)
        if col_index is None:
            return None
        return row_data.get(col_index)

    # ================================================================
    # Acceso a datos
    # ================================================================

    def get_articles(self) -> dict[str, dict[str, Any]]:
        """
        Retorna los datos de artículos cargados del Excel.

        Returns:
            Diccionario {código: {"description": str, "quantity": int}}.
            Diccionario vacío si no hay datos cargados.
        """
        return self._articles_data.copy()

    def get_article_data(self, code: str) -> Optional[dict[str, Any]]:
        """
        Obtiene los datos de un artículo específico por código.

        Args:
            code: Código del artículo.

        Returns:
            Diccionario con description y quantity, o None si no existe.
        """
        return self._articles_data.get(code.upper())

    # ================================================================
    # Fusión con artículos escaneados
    # ================================================================

    def merge_with_scanned(self, articles: list[Article]) -> list[Article]:
        """
        Enriquece artículos escaneados con datos del Excel.

        Agrega la descripción del Excel a cada artículo que coincida
        por código. Los artículos sin coincidencia mantienen su
        descripción vacía.

        Args:
            articles: Lista de artículos escaneados desde imágenes.

        Returns:
            La misma lista de artículos con descripciones actualizadas.
        """
        if not self.has_data:
            logger.debug("Sin datos Excel para fusionar.")
            return articles

        merged_count = 0
        for article in articles:
            excel_data = self._articles_data.get(article.code)
            if excel_data is not None:
                article.description = excel_data["description"]
                merged_count += 1

        logger.info(
            f"Fusión Excel completada: {merged_count}/{len(articles)} "
            f"artículos enriquecidos con descripción"
        )
        return articles

    def __repr__(self) -> str:
        if self._file_path:
            return (
                f"ExcelService(file='{self._file_path.name}', "
                f"loaded={self._loaded}, articles={len(self._articles_data)})"
            )
        return "ExcelService(sin archivo)"
