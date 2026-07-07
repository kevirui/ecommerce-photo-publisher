"""
Servicio de generación de reportes Excel de publicación.

Genera un archivo Reporte_Publicacion.xlsx en el Escritorio del usuario
con el resultado de cada artículo: código, estado, cantidad de imágenes,
tiempo transcurrido y mensaje de error. Aplica formato condicional
por estado (verde/rojo/amarillo).
"""

from __future__ import annotations

import logging
from datetime import datetime
from pathlib import Path
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from models.article import Article, ArticleStatus

logger = logging.getLogger(__name__)

# ============================================================
# Colores de formato condicional por estado
# ============================================================

FILL_SUCCESS = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
FILL_ERROR = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
FILL_WARNING = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
FILL_HEADER = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
FILL_SKIPPED = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

FONT_HEADER = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
FONT_NORMAL = Font(name="Segoe UI", size=10)
FONT_SUCCESS = Font(name="Segoe UI", size=10, color="006100")
FONT_ERROR = Font(name="Segoe UI", size=10, color="9C0006")
FONT_WARNING = Font(name="Segoe UI", size=10, color="9C5700")
FONT_TITLE = Font(name="Segoe UI", size=14, bold=True, color="1F4E79")

BORDER_THIN = Border(
    left=Side(style="thin", color="B4B4B4"),
    right=Side(style="thin", color="B4B4B4"),
    top=Side(style="thin", color="B4B4B4"),
    bottom=Side(style="thin", color="B4B4B4"),
)

# Columnas del reporte
REPORT_COLUMNS = [
    ("Código", 18),
    ("Estado", 15),
    ("Cant. Imágenes", 16),
    ("Tiempo", 12),
    ("Mensaje", 60),
]


class ReportServiceError(Exception):
    """Excepción base para errores del servicio de reportes."""
    pass


class ReportService:
    """
    Servicio para generar reportes Excel de publicación.

    Genera un archivo Excel formateado con el resultado de cada artículo
    procesado: código, estado, cantidad de imágenes, tiempo y mensaje.
    El reporte se guarda por defecto en el Escritorio del usuario.

    Example:
        >>> report = ReportService()
        >>> path = report.generate_report(articles)
        >>> print(f"Reporte guardado en: {path}")
    """

    @staticmethod
    def _get_desktop_path() -> Path:
        """
        Obtiene la ruta al Escritorio del usuario.

        Returns:
            Path al directorio del Escritorio.
        """
        desktop = Path.home() / "Desktop"
        if not desktop.exists():
            # Fallback para sistemas en español
            desktop_es = Path.home() / "Escritorio"
            if desktop_es.exists():
                return desktop_es
            # Crear Desktop si no existe
            desktop.mkdir(exist_ok=True)
        return desktop

    def generate_report(
        self,
        articles: list[Article],
        output_path: Optional[Path] = None,
    ) -> Path:
        """
        Genera el reporte Excel de publicación.

        Args:
            articles: Lista de artículos procesados.
            output_path: Ruta personalizada para el reporte.
                        Si es None, se guarda en el Escritorio.

        Returns:
            Ruta al archivo Excel generado.

        Raises:
            ReportServiceError: Si ocurre un error al generar el reporte.
        """
        if output_path is None:
            desktop = self._get_desktop_path()
            output_path = desktop / self._generate_filename()

        try:
            logger.info(f"Generando reporte de publicación: {output_path}")

            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Reporte Publicación"

            # --- Título ---
            self._write_title(sheet, articles)

            # --- Encabezados ---
            header_row = 3
            self._write_headers(sheet, header_row)

            # --- Datos ---
            self._write_data(sheet, articles, header_row + 1)

            # --- Ajustar anchos ---
            self._adjust_column_widths(sheet)

            # --- Resumen al final ---
            summary_row = header_row + len(articles) + 2
            self._write_summary(sheet, articles, summary_row)

            # Guardar
            workbook.save(str(output_path))
            workbook.close()

            logger.info(
                f"Reporte generado exitosamente: {output_path} "
                f"({len(articles)} artículos)"
            )
            return output_path

        except Exception as e:
            logger.error(f"Error al generar reporte: {e}", exc_info=True)
            raise ReportServiceError(f"Error al generar reporte: {e}")

    # ================================================================
    # Escritura de secciones
    # ================================================================

    @staticmethod
    def _write_title(sheet, articles: list[Article]) -> None:
        """Escribe el título del reporte en la primera fila."""
        timestamp = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
        sheet.merge_cells("A1:E1")
        title_cell = sheet["A1"]
        title_cell.value = f"Reporte de Publicación — {timestamp}"
        title_cell.font = FONT_TITLE
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        sheet.row_dimensions[1].height = 30

    @staticmethod
    def _write_headers(sheet, row: int) -> None:
        """Escribe los encabezados de columna."""
        for col_idx, (col_name, _) in enumerate(REPORT_COLUMNS, start=1):
            cell = sheet.cell(row=row, column=col_idx, value=col_name)
            cell.font = FONT_HEADER
            cell.fill = FILL_HEADER
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = BORDER_THIN
        sheet.row_dimensions[row].height = 25

    @staticmethod
    def _write_data(sheet, articles: list[Article], start_row: int) -> None:
        """Escribe los datos de cada artículo."""
        for idx, article in enumerate(articles):
            row = start_row + idx

            # Determinar estilo según estado
            fill, font = ReportService._get_status_style(article.status)

            # Código
            cell_code = sheet.cell(row=row, column=1, value=article.code)
            cell_code.font = font
            cell_code.fill = fill
            cell_code.border = BORDER_THIN

            # Estado
            cell_status = sheet.cell(
                row=row, column=2, value=str(article.status)
            )
            cell_status.font = font
            cell_status.fill = fill
            cell_status.alignment = Alignment(horizontal="center")
            cell_status.border = BORDER_THIN

            # Cantidad de imágenes
            cell_qty = sheet.cell(
                row=row, column=3, value=article.image_count
            )
            cell_qty.font = font
            cell_qty.fill = fill
            cell_qty.alignment = Alignment(horizontal="center")
            cell_qty.border = BORDER_THIN

            # Tiempo
            cell_time = sheet.cell(
                row=row, column=4, value=article.elapsed_time_formatted
            )
            cell_time.font = font
            cell_time.fill = fill
            cell_time.alignment = Alignment(horizontal="center")
            cell_time.border = BORDER_THIN

            # Mensaje
            message = article.error_message if article.error_message else "OK"
            cell_msg = sheet.cell(row=row, column=5, value=message)
            cell_msg.font = font
            cell_msg.fill = fill
            cell_msg.border = BORDER_THIN

    @staticmethod
    def _get_status_style(status: ArticleStatus) -> tuple[PatternFill, Font]:
        """
        Retorna el estilo de celda según el estado del artículo.

        Args:
            status: Estado del artículo.

        Returns:
            Tupla (fill, font) con los estilos apropiados.
        """
        styles = {
            ArticleStatus.SUCCESS: (FILL_SUCCESS, FONT_SUCCESS),
            ArticleStatus.ERROR: (FILL_ERROR, FONT_ERROR),
            ArticleStatus.PENDING: (FILL_WARNING, FONT_WARNING),
            ArticleStatus.IN_PROGRESS: (FILL_WARNING, FONT_WARNING),
            ArticleStatus.SKIPPED: (FILL_SKIPPED, FONT_NORMAL),
        }
        return styles.get(status, (FILL_WARNING, FONT_NORMAL))

    @staticmethod
    def _write_summary(sheet, articles: list[Article], row: int) -> None:
        """Escribe el resumen al final del reporte."""
        total = len(articles)
        success = sum(1 for a in articles if a.status == ArticleStatus.SUCCESS)
        errors = sum(1 for a in articles if a.status == ArticleStatus.ERROR)
        skipped = sum(1 for a in articles if a.status == ArticleStatus.SKIPPED)
        pending = sum(1 for a in articles if a.status == ArticleStatus.PENDING)

        total_time = sum(a.elapsed_time for a in articles)
        time_str = (
            f"{int(total_time // 60)}m {total_time % 60:.0f}s"
            if total_time >= 60
            else f"{total_time:.1f}s"
        )

        summary_lines = [
            ("Total artículos:", str(total)),
            ("Exitosos:", str(success)),
            ("Con error:", str(errors)),
            ("Omitidos:", str(skipped)),
            ("Pendientes:", str(pending)),
            ("Tiempo total:", time_str),
        ]

        font_bold = Font(name="Segoe UI", size=10, bold=True)

        for idx, (label, value) in enumerate(summary_lines):
            label_cell = sheet.cell(row=row + idx, column=1, value=label)
            label_cell.font = font_bold

            value_cell = sheet.cell(row=row + idx, column=2, value=value)
            value_cell.font = FONT_NORMAL

    @staticmethod
    def _adjust_column_widths(sheet) -> None:
        """Ajusta el ancho de las columnas según la configuración."""
        for col_idx, (_, width) in enumerate(REPORT_COLUMNS, start=1):
            col_letter = get_column_letter(col_idx)
            sheet.column_dimensions[col_letter].width = width

    @staticmethod
    def _generate_filename() -> str:
        """
        Genera un nombre de archivo único para el reporte.

        Returns:
            Nombre del archivo con timestamp.
        """
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        return f"Reporte_Publicacion_{timestamp}.xlsx"

    def __repr__(self) -> str:
        return "ReportService()"
