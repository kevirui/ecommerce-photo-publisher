"""
Servicio de Sincronización y Comparación de Artículos.

Se encarga de:
1. Leer el archivo Excel e identificar automáticamente la columna de código.
2. Comparar los códigos contra SQL Server usando SqlService.
3. Determinar los estados (PUBLICADO, PENDIENTE, INCOMPLETO, INEXISTENTE).
4. Administrar un caché de sesión para evitar consultas duplicadas.
5. Exportar los resultados filtrados y generar el reporte automático 'Pendientes.xlsx'.
"""

import logging
import unicodedata
from pathlib import Path
from typing import Any, Optional, Dict, List, Tuple
from openpyxl import load_workbook, Workbook

from services.sql_service import SqlService

logger = logging.getLogger(__name__)


def normalize_header(text: str) -> str:
    """Normaliza texto eliminando acentos y convirtiendo a minúsculas."""
    if not text:
        return ""
    normalized = "".join(
        c for c in unicodedata.normalize("NFD", text)
        if unicodedata.category(c) != "Mn"
    )
    return normalized.lower().strip()


class SyncService:
    """
    Servicio de negocio para la sincronización y comparación de artículos.
    """

    def __init__(self, sql_service: SqlService) -> None:
        """
        Inicializa el servicio de sincronización.

        Args:
            sql_service: Instancia activa de SqlService.
        """
        self._sql_service = sql_service
        self._cache: Dict[str, Dict[str, Any]] = {}

    def clear_cache(self) -> None:
        """Limpia el caché de artículos de la sesión."""
        self._cache.clear()
        logger.info("Caché de Sincronización limpiado.")

    def read_excel_codes(self, file_path: Path) -> List[str]:
        """
        Lee un archivo Excel (tanto .xlsx como .xls antiguo) y retorna la lista de códigos de artículos
        detectados en la columna correspondiente.
        """
        if not file_path.exists():
            raise FileNotFoundError(f"El archivo Excel no existe: {file_path}")

        suffix = file_path.suffix.lower()
        if suffix == ".xls":
            return self._read_xls_codes(file_path)
        else:
            return self._read_xlsx_codes(file_path)

    def _read_xls_codes(self, file_path: Path) -> List[str]:
        """Lee códigos desde un archivo .xls usando la biblioteca xlrd."""
        try:
            import xlrd
        except ImportError:
            raise ImportError(
                "El formato de archivo .xls requiere el paquete 'xlrd'.\n"
                "Instálelo ejecutando 'pip install xlrd' o guarde su archivo como .xlsx (Libro de Excel moderno)."
            )

        logger.info(f"Leyendo Excel legacy (.xls) para códigos: {file_path}")
        workbook = xlrd.open_workbook(filename=str(file_path))
        sheet = workbook.sheet_by_index(0)

        if sheet.nrows == 0:
            raise ValueError("El archivo Excel legacy .xls no contiene filas.")

        # Detectar columna de código en la primera fila
        first_row = [sheet.cell_value(0, col) for col in range(sheet.ncols)]
        code_col_idx: Optional[int] = None
        accepted_names = {"codigo", "articulo", "cod_articulo"}

        for idx, val in enumerate(first_row):
            if val is None:
                continue
            normalized_val = normalize_header(str(val))
            normalized_val_clean = normalized_val.replace("_", "").replace("-", "")
            if normalized_val in accepted_names or normalized_val_clean in accepted_names:
                code_col_idx = idx
                logger.info(f"Columna de código detectada: '{val}' en columna {code_col_idx}")
                break

        if code_col_idx is None:
            raise ValueError(
                f"No se pudo detectar automáticamente la columna del código en el archivo .xls.\n"
                f"Asegúrese de tener una columna llamada 'Código', 'Codigo', 'Articulo', etc.\n"
                f"Columnas encontradas: {first_row}"
            )

        codes: List[str] = []
        for row_idx in range(1, sheet.nrows):
            cell_val = sheet.cell_value(row_idx, code_col_idx)
            if cell_val is not None:
                # xlrd lee enteros como números flotantes (.0), los formateamos a entero si corresponde
                if isinstance(cell_val, float) and cell_val.is_integer():
                    code_str = str(int(cell_val)).strip()
                else:
                    code_str = str(cell_val).strip()
                
                if code_str:
                    if code_str not in codes:
                        codes.append(code_str)

        logger.info(f"Se leyeron {len(codes)} códigos del archivo Excel legacy .xls.")
        return codes

    def _read_xlsx_codes(self, file_path: Path) -> List[str]:
        """Lee códigos desde un archivo .xlsx usando openpyxl."""
        logger.info(f"Leyendo Excel (.xlsx) para códigos: {file_path}")
        workbook = load_workbook(filename=str(file_path), read_only=True, data_only=True)
        sheet = workbook.active

        if sheet is None:
            workbook.close()
            raise ValueError("El archivo Excel no contiene hojas activas.")

        first_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=False))
        code_col_idx: Optional[int] = None
        accepted_names = {"codigo", "articulo", "cod_articulo"}

        for cell in first_row:
            if cell.value is None:
                continue
            normalized_val = normalize_header(str(cell.value))
            normalized_val_clean = normalized_val.replace("_", "").replace("-", "")
            if normalized_val in accepted_names or normalized_val_clean in accepted_names:
                code_col_idx = cell.column
                logger.info(f"Columna de código detectada: '{cell.value}' en columna {code_col_idx}")
                break

        if code_col_idx is None:
            found_headers = [str(c.value) for c in first_row if c.value is not None]
            workbook.close()
            raise ValueError(
                f"No se pudo detectar automáticamente la columna del código.\n"
                f"Asegúrese de tener una columna llamada 'Código', 'Codigo', 'Articulo', etc.\n"
                f"Columnas encontradas: {found_headers}"
            )

        codes: List[str] = []
        for row in sheet.iter_rows(min_row=2, values_only=False):
            cell_val = None
            for cell in row:
                if cell.column == code_col_idx:
                    cell_val = cell.value
                    break

            if cell_val is not None:
                # openpyxl puede leer números como float o int
                if isinstance(cell_val, float) and cell_val.is_integer():
                    code_str = str(int(cell_val)).strip()
                else:
                    code_str = str(cell_val).strip()

                if code_str:
                    if code_str not in codes:
                        codes.append(code_str)

        workbook.close()
        logger.info(f"Se leyeron {len(codes)} códigos del archivo Excel.")
        return codes

    def get_article_status_from_db(self, code: str) -> Dict[str, Any]:
        """
        Consulta SQL Server por un artículo específico y retorna su información básica,
        utilizando caché de sesión si está disponible.

        Args:
            code: Código del artículo.
        """
        code_upper = code.strip().upper()

        if code_upper in self._cache:
            logger.debug(f"Caché hit para artículo: {code_upper}")
            return self._cache[code_upper]

        query = """
        SELECT
            A.COD_ARTICULO,
            A.DESCRIP_ARTI,
            A.WEB_PUBLI,
            A.WEB_IMAGEN_PROVE,
            A.CANT_STOCK,
            COUNT(I.COD_ARTICULO)
        FROM ARTICULOS A
        LEFT JOIN ARTICULOS_IMAGENES I
            ON A.COD_ARTICULO = I.COD_ARTICULO
        WHERE A.COD_ARTICULO = ?
        GROUP BY
            A.COD_ARTICULO,
            A.DESCRIP_ARTI,
            A.WEB_PUBLI,
            A.WEB_IMAGEN_PROVE,
            A.CANT_STOCK
        """

        try:
            results = self._sql_service.execute(query, (code_upper,))
            if not results:
                art_data = {
                    "codigo": code_upper,
                    "descripcion": "No encontrado en base de datos",
                    "publicado_db": "N",
                    "imagen_principal": "",
                    "cant_imagenes": 0,
                    "cant_stock": 0.0,
                    "estado": "INEXISTENTE",
                    "observaciones": "El artículo no existe en la tabla ARTICULOS.",
                    "web_link": ""
                }
            else:
                row = results[0]
                descrip = row.get("DESCRIP_ARTI", "")
                web_publi = row.get("WEB_PUBLI", "")
                imagen_prov = row.get("WEB_IMAGEN_PROVE", "")
                cant_stock = row.get("CANT_STOCK", 0.0)
                
                known_keys = {"COD_ARTICULO", "DESCRIP_ARTI", "WEB_PUBLI", "WEB_IMAGEN_PROVE", "CANT_STOCK"}
                count_key = next((k for k in row.keys() if k not in known_keys), None)
                cant_imagenes = int(row[count_key]) if count_key is not None and row[count_key] is not None else 0

                web_link = ""

                is_publi = (web_publi == "S")
                has_any_image = (imagen_prov is not None and str(imagen_prov).strip() != "") or (cant_imagenes > 0)

                if not is_publi:
                    estado = "PENDIENTE"
                    observaciones = "El artículo no está marcado para publicar en la web."
                elif not has_any_image:
                    estado = "INCOMPLETO"
                    observaciones = "Marcado para publicar pero no tiene ninguna imagen (principal o adicional)."
                else:
                    estado = "PUBLICADO"
                    observaciones = "Artículo publicado (tiene al menos una imagen)."

                art_data = {
                    "codigo": code_upper,
                    "descripcion": descrip if descrip else "",
                    "publicado_db": web_publi if web_publi else "N",
                    "imagen_principal": imagen_prov if imagen_prov else "",
                    "cant_imagenes": cant_imagenes,
                    "cant_stock": float(cant_stock) if cant_stock is not None else 0.0,
                    "estado": estado,
                    "observaciones": observaciones,
                    "web_link": web_link
                }

            self._cache[code_upper] = art_data
            return art_data

        except Exception as e:
            logger.error(f"Error consultando artículo {code_upper}: {e}")
            raise

    def get_all_articles_status_from_db(self, pending_only: bool = False) -> List[Dict[str, Any]]:
        """
        Consulta todos los artículos directamente de la base de datos (Modo Base de Datos).
        """
        query = """
        SELECT
            A.COD_ARTICULO,
            A.DESCRIP_ARTI,
            A.WEB_PUBLI,
            A.WEB_IMAGEN_PROVE,
            A.CANT_STOCK,
            COUNT(I.COD_ARTICULO)
        FROM ARTICULOS A
        LEFT JOIN ARTICULOS_IMAGENES I
            ON A.COD_ARTICULO = I.COD_ARTICULO
        """

        if pending_only:
            query += " WHERE A.WEB_PUBLI <> 'S' OR A.WEB_IMAGEN_PROVE IS NULL OR A.WEB_IMAGEN_PROVE = '' "

        query += """
        GROUP BY
            A.COD_ARTICULO,
            A.DESCRIP_ARTI,
            A.WEB_PUBLI,
            A.WEB_IMAGEN_PROVE,
            A.CANT_STOCK
        """

        logger.info(f"Ejecutando consulta SQL general (pending_only={pending_only})")
        results = self._sql_service.execute(query)
        articles: List[Dict[str, Any]] = []

        for row in results:
            code = row.get("COD_ARTICULO", "")
            descrip = row.get("DESCRIP_ARTI", "")
            web_publi = row.get("WEB_PUBLI", "")
            imagen_prov = row.get("WEB_IMAGEN_PROVE", "")
            cant_stock = row.get("CANT_STOCK", 0.0)

            known_keys = {"COD_ARTICULO", "DESCRIP_ARTI", "WEB_PUBLI", "WEB_IMAGEN_PROVE", "CANT_STOCK"}
            count_key = next((k for k in row.keys() if k not in known_keys), None)
            cant_imagenes = int(row[count_key]) if count_key is not None and row[count_key] is not None else 0

            is_publi = (web_publi == "S")
            has_any_image = (imagen_prov is not None and str(imagen_prov).strip() != "") or (cant_imagenes > 0)

            # Si se solicita solo pendientes de fotografiar, y ya tiene alguna imagen y está publicado, omitir
            if pending_only and is_publi and has_any_image:
                continue

            if not is_publi:
                estado = "PENDIENTE"
                observaciones = "El artículo no está marcado para publicar en la web."
            elif not has_any_image:
                estado = "INCOMPLETO"
                observaciones = "Marcado para publicar pero no tiene ninguna imagen (principal o adicional)."
            else:
                estado = "PUBLICADO"
                observaciones = "Artículo publicado (tiene al menos una imagen)."

            art_data = {
                "codigo": code,
                "descripcion": descrip if descrip else "",
                "publicado_db": web_publi if web_publi else "N",
                "imagen_principal": imagen_prov if imagen_prov else "",
                "cant_imagenes": cant_imagenes,
                "cant_stock": float(cant_stock) if cant_stock is not None else 0.0,
                "estado": estado,
                "observaciones": observaciones,
                "web_link": ""
            }

            self._cache[code] = art_data
            articles.append(art_data)

        return articles

    def query_detailed_article(self, code: str) -> Dict[str, Any]:
        """
        Obtiene información detallada de un artículo para la ventana de doble click.
        """
        query = "SELECT WEB_LINK, WEB_IMAGEN_PROVE FROM ARTICULOS WHERE COD_ARTICULO = ?"
        try:
            res = self._sql_service.execute(query, (code,))
            web_link = ""
            web_img = ""
            if res:
                web_link = res[0].get("WEB_LINK", "")
                web_img = res[0].get("WEB_IMAGEN_PROVE", "")
                if code in self._cache:
                    self._cache[code]["web_link"] = web_link if web_link else ""
                    if web_img:
                        self._cache[code]["imagen_principal"] = web_img
            return {
                "web_link": web_link if web_link else "",
                "imagen_principal": web_img if web_img else ""
            }
        except Exception as e:
            logger.error(f"Error consultando detalle extra para {code}: {e}")
            return {"web_link": "", "imagen_principal": ""}

    def export_to_excel(self, file_path: Path, data: List[Dict[str, Any]], filter_type: str = "Todos") -> None:
        """
        Exporta los datos proporcionados a un archivo Excel.
        """
        logger.info(f"Exportando artículos ({filter_type}) a: {file_path}")
        wb = Workbook()
        ws = wb.active
        ws.title = "Artículos"

        headers = ["Código", "Descripción", "Publicado en Web", "Imagen Principal", "Cantidad Imágenes", "Estado", "Observaciones"]
        ws.append(headers)

        for item in data:
            ws.append([
                item.get("codigo", ""),
                item.get("descripcion", ""),
                item.get("publicado_db", ""),
                item.get("imagen_principal", ""),
                item.get("cant_imagenes", 0),
                item.get("estado", ""),
                item.get("observaciones", "")
            ])

        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = col[0].column_letter
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

        wb.save(str(file_path))
        wb.close()
        logger.info(f"Exportación completada: {file_path}")

    def export_pending_auto(self, output_dir: Path, data: List[Dict[str, Any]]) -> Path:
        """
        Genera automáticamente el archivo 'Pendientes.xlsx' con artículos PENDIENTE que tengan stock.
        """
        file_path = output_dir / "Pendientes.xlsx"
        logger.info(f"Generando automáticamente el reporte de pendientes: {file_path}")

        wb = Workbook()
        ws = wb.active
        ws.title = "Pendientes"

        headers = ["Código", "Descripción", "Estado", "Stock", "Cantidad imágenes", "Observaciones"]
        ws.append(headers)

        count = 0
        for item in data:
            if item.get("estado") == "PENDIENTE":
                cant_stock = float(item.get("cant_stock", 0.0))
                if cant_stock > 0:
                    ws.append([
                        item.get("codigo", ""),
                        item.get("descripcion", ""),
                        item.get("estado", ""),
                        cant_stock,
                        item.get("cant_imagenes", 0),
                        item.get("observaciones", "")
                    ])
                    count += 1

        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = col[0].column_letter
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

        wb.save(str(file_path))
        wb.close()
        logger.info(f"Reporte de pendientes generado ({count} artículos) en: {file_path}")
        return file_path
